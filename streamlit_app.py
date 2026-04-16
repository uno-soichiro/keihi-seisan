import streamlit as st
import anthropic
import base64
import json
import os
import io
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from copy import copy
import fitz  # PyMuPDF
# ─────────────────────────────────────────────
# ページ設定
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="経費精算ツール",
    page_icon="📊",
    layout="centered",
)

st.markdown("""
<style>
.main { max-width: 720px; }
.stButton > button {
    width: 100%;
    background: linear-gradient(135deg, #4f7ef7, #7c4dff);
    color: white;
    font-weight: 600;
    font-size: 16px;
    padding: 12px;
    border-radius: 10px;
    border: none;
}
.stDownloadButton > button {
    width: 100%;
    background: #22c55e;
    color: white;
    font-weight: 600;
    font-size: 16px;
    padding: 12px;
    border-radius: 10px;
    border: none;
}
.result-box {
    background: #1e1e2e;
    color: #a9b1d6;
    border-radius: 10px;
    padding: 16px;
    font-family: monospace;
    font-size: 13px;
    white-space: pre-wrap;
}
</style>
""", unsafe_allow_html=True)

TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"
MEDIA_TYPES = {
    '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
    '.gif': 'image/gif', '.webp': 'image/webp', '.pdf': 'application/pdf',
}


# ─────────────────────────────────────────────
# Claude API で領収書を解析
# ─────────────────────────────────────────────
def extract_receipt_data(client, file_bytes: bytes, filename: str) -> list:
    ext = Path(filename).suffix.lower()
    media_type = MEDIA_TYPES.get(ext, 'image/jpeg')
    img_data = base64.standard_b64encode(file_bytes).decode('utf-8')

    prompt = """この画像に含まれる全ての領収書・レシートから経費情報を抽出してください。

【重要】1枚の画像に複数の領収書・レシートが含まれている場合は、それぞれを個別に抽出してください。
領収書が1枚だけの場合も含め、必ずJSON配列形式で返してください。

以下のJSON配列形式のみを返してください（コードブロック不要）:
[{"type":"travel","date":"YYYY-MM-DD","amount":金額,"category":"科目","description":"摘要","route":"経路またはnull"}]

typeの選び方:
- 電車・新幹線・バス・タクシーなど交通系 → "travel"
- 接待・会食・備品・宿泊など → "other"

categoryの例:
- travel: 「旅費」(新幹線・特急など)、「交通費」(電車・バス・タクシー)
- other: 「接待交際費」「会議費」「消耗品費」「宿泊費」「諸経費」

routeは交通費のみ（例:品川→熱海）。その他はnull。
日付不明はnull、金額不明は0。
品目・説明が不明な場合はdescriptionを「不明」とする。科目が判断できない場合はcategoryを「諸経費」とする。
JSON配列のみ返してください。"""

    # PDFは最初のページを画像に変換してから送信
    if ext == '.pdf':
        pdf_doc = fitz.open(stream=file_bytes, filetype="pdf")
        page = pdf_doc[0]  # 1ページ目
        mat = fitz.Matrix(1.5, 1.5)  # サイズ削減のため解像度を下げる
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("jpeg", jpg_quality=85)  # PNGよりJPEGの方が大幅に小さい
        img_data = base64.standard_b64encode(img_bytes).decode('utf-8')
        content_block = {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": img_data}}
        pdf_doc.close()
    else:
        content_block = {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_data}}

    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=512,
        messages=[{"role": "user", "content": [
            content_block,
            {"type": "text", "text": prompt}
        ]}]
    )

    text = resp.content[0].text.strip()
    if '```' in text:
        for part in text.split('```'):
            part = part.strip()
            if part.startswith('json'):
                part = part[4:].strip()
            try:
                result = json.loads(part)
                return result if isinstance(result, list) else [result]
            except Exception:
                continue
    result = json.loads(text)
    return result if isinstance(result, list) else [result]


# ─────────────────────────────────────────────
# Excel生成
# ─────────────────────────────────────────────
def write_sheet(ws, items: list, data_start: int, data_end: int, has_route: bool, count_cell: str = None):
    import re
    original_data_end = data_end

    # 項目数がテンプレートの行数を超える場合、合計行の前に行を挿入してずらす
    available = data_end - data_start + 1
    if len(items) > available:
        extra = len(items) - available
        ws.insert_rows(data_end + 1, extra)
        data_end += extra

    # データ行をクリア
    for r in range(data_start, data_end + 1):
        for c in range(1, 12):
            ws.cell(row=r, column=c).value = None

    # データ書き込み
    for i, item in enumerate(items):
        row = data_start + i
        ws.cell(row=row, column=1).value = f'=ROW()-{data_start - 1}'
        if item.get('date'):
            try:
                ws.cell(row=row, column=2).value = datetime.strptime(item['date'], '%Y-%m-%d')
                ws.cell(row=row, column=2).number_format = 'm/d'
            except Exception:
                ws.cell(row=row, column=2).value = item['date']
        amt = item.get('amount', 0)
        ws.cell(row=row, column=3).value = amt if amt else None
        ws.cell(row=row, column=3).number_format = '[$¥-411]#,##0'
        ws.cell(row=row, column=4).value = item.get('category', '')
        ws.cell(row=row, column=5).value = item.get('description', '')
        if has_route:
            ws.cell(row=row, column=8).value = item.get('route') or ''

    # 行挿入でSUM関数の範囲がズレた場合、新しいdata_endまで拡張する
    if data_end > original_data_end:
        pattern = re.compile(
            r'([A-Za-z]+)' + str(data_start) + r':([A-Za-z]+)' + str(original_data_end),
        )
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value and isinstance(cell.value, str) and 'SUM' in cell.value.upper():
                    new_val = pattern.sub(
                        lambda m: f'{m.group(1)}{data_start}:{m.group(2)}{data_end}',
                        cell.value
                    )
                    if new_val != cell.value:
                        cell.value = new_val

    # 件数セルへの書き込み（シートごとに指定）
    if count_cell:
        ws[count_cell] = len(items)

def generate_excel(travel_items: list, other_items: list, person_name: str = '') -> bytes:
    wb = load_workbook(str(TEMPLATE_PATH))
    ws_t = wb['旅費交通費']
    ws_o = wb['その他']

    write_sheet(ws_t, travel_items, 13, 26, True, count_cell='D8')
    write_sheet(ws_o, other_items, 13, 19, False, count_cell=None)

    today = datetime.now()
    ws_t['D4'] = today
    ws_o['D4'] = today

    if person_name:
        ws_t['D6'] = person_name
        ws_o['D6'] = person_name

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────
st.title("📊 経費精算ツール")
st.caption("領収書をアップロードするだけで、Excelの経費精算書を自動作成します")

st.divider()

# APIキー入力
st.subheader("🔑 Anthropic APIキー")

# 環境変数（Streamlit Cloudでの secrets 対応）
env_key = os.environ.get('ANTHROPIC_API_KEY', '') or st.secrets.get('ANTHROPIC_API_KEY', '') if hasattr(st, 'secrets') else ''
if env_key:
    api_key = env_key
    st.success("✅ APIキーが設定済みです（環境変数）")
else:
    api_key = st.text_input(
        "APIキーを入力してください",
        type="password",
        placeholder="sk-ant-...",
        help="[Anthropic Console](https://console.anthropic.com) でAPIキーを取得できます"
    )

st.divider()

# 精算者選択
st.subheader("👤 精算者")
person_name = st.selectbox(
    "精算者を選択してください",
    options=["小田崇", "宇野想一郎"],
    index=0,
)

st.divider()

# ファイルアップロード
st.subheader("📎 領収書をアップロード")
uploaded_files = st.file_uploader(
    "領収書画像・PDFをアップロード（複数選択可）",
    type=['jpg', 'jpeg', 'png', 'gif', 'webp', 'pdf'],
    accept_multiple_files=True,
    help="対応形式: JPG, PNG, PDF, WEBP, GIF"
)

# ファイルサイズ上限（バイト単位） - 10MB
MAX_FILE_SIZE_MB = 10
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

if uploaded_files:
    st.write(f"📁 **{len(uploaded_files)} ファイル**が選択されています:")
    cols = st.columns(min(len(uploaded_files), 4))
    oversized_files = []
    for i, f in enumerate(uploaded_files):
        with cols[i % 4]:
            ext = Path(f.name).suffix.lower()
            file_size = f.size
            if file_size > MAX_FILE_SIZE_BYTES:
                oversized_files.append((f.name, file_size))
                st.error(f"⚠️ `{f.name}`\nサイズ超過 ({file_size / 1024 / 1024:.1f}MB) — {MAX_FILE_SIZE_MB}MB以下にしてください")
            elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                st.image(f, caption=f.name, use_container_width=True)
            else:
                st.markdown(f"📄 `{f.name}`")
    if oversized_files:
        st.warning(f"🚫 {len(oversized_files)}件のファイルがサイズ上限({MAX_FILE_SIZE_MB}MB)を超えています。処理を実行する前に、これらのファイルを削除してください。")

st.divider()

# 処理ボタン
oversized_files = [f for f in (uploaded_files or []) if f.size > MAX_FILE_SIZE_BYTES]
if st.button("⚡ Excel経費精算書を作成する", disabled=not (api_key and uploaded_files) or bool(oversized_files)):

    if not TEMPLATE_PATH.exists():
        st.error("テンプレートファイル (template.xlsx) が見つかりません。")
        st.stop()

    try:
        client = anthropic.Anthropic(api_key=api_key)
    except Exception as e:
        st.error(f"APIキーが無効です: {e}")
        st.stop()

    travel_items = []
    other_items = []
    log_lines = []

    progress_bar = st.progress(0)
    status_text = st.empty()

    for i, uploaded_file in enumerate(uploaded_files):
        status_text.text(f"🔍 解析中... {uploaded_file.name} ({i+1}/{len(uploaded_files)})")
        progress_bar.progress((i) / len(uploaded_files))

        try:
            file_bytes = uploaded_file.read()
            items = extract_receipt_data(client, file_bytes, uploaded_file.name)
            for data in items:
                amt = data.get('amount', 0)
                amt_str = f"¥{int(amt):,}" if isinstance(amt, (int, float)) and amt else "¥0"
                log_line = f"✅ {uploaded_file.name}\n   → {data.get('date', '日付不明')} | {data.get('category', '不明')} | {amt_str}"
                if data.get('route'):
                    log_line += f" | {data.get('route')}"
                log_lines.append(log_line)
                if data.get('type') == 'travel':
                    travel_items.append(data)
                else:
                    other_items.append(data)
        except json.JSONDecodeError as e:
            log_lines.append(f"⚠️ {uploaded_file.name}\n   → JSON解析エラー: {e}")
        except Exception as e:
            full_err = str(e)
            if hasattr(e, 'body') and e.body:
                full_err = str(e.body)
            elif hasattr(e, 'message') and e.message:
                full_err = e.message
            import traceback, sys
            print(f"[FULL_ERROR] {uploaded_file.name}: {repr(full_err)}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            log_lines.append(f"❌ {uploaded_file.name}\n   → {full_err}")

    progress_bar.progress(1.0)
    status_text.text("✨ Excel生成中...")

    # 日付ソート
    travel_items.sort(key=lambda x: x.get('date') or '9999-99-99')
    other_items.sort(key=lambda x: x.get('date') or '9999-99-99')

    try:
        excel_bytes = generate_excel(travel_items, other_items, person_name)
        status_text.empty()
        progress_bar.empty()

        # 結果表示
        st.success(f"✅ 完了！ 旅費交通費: **{len(travel_items)}件** / その他: **{len(other_items)}件**")

        # ログ
        with st.expander("📋 処理ログを見る", expanded=True):
            for line in log_lines:
                if line.startswith('❌'):
                    st.error(line)
                elif line.startswith('⚠️'):
                    st.warning(line)
                else:
                    st.success(line)

        # ダウンロードボタン
        today_str = datetime.now().strftime('%Y%m')
        st.download_button(
            label="📥 Excelファイルをダウンロード",
            data=excel_bytes,
            file_name=f"{today_str}経費精算.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        status_text.empty()
        st.error(f"Excel生成エラー: {e}")

st.divider()
st.markdown(
    '<p style="text-align:center; color:#aaa; font-size:12px;">Powered by Claude AI · 領収書の情報は外部に保存されません</p>',
    unsafe_allow_html=True
)
